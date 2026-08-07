#!/usr/bin/env python3
"""Load exCALIBR MAVE calibrations into Supplementary Data 4.

Reads every `*.json` calibration file produced by exCALIBR (default:
`data/input/mave_calibration/excalibr/exc_pp_calib_final_fixedmapping_clinvar/json/`)
and writes one row per file into the `ExCALIBR_calibrations` sheet of a
workbook (default: `data/output/supplementary_data/Supplementary_Data_4.xlsx`),
overwriting that sheet's existing contents in place. Other sheets are left
untouched.

Each file's name (minus `.json`) becomes the `dataset` column. `point_ranges`
supplies the sixteen `range_-8` .. `range_8` columns: a point with no range is
left blank, and a populated range is written as its two endpoints separated
by a space, with `-Infinity`/`Infinity` rendered as `-inf`/`inf`. `relax`,
`clinvar_2018`, and `scoreset_flipped` are 0/1 in the source JSON and are
written as Excel booleans; `prior`, `n_c`, and `benign_method` are copied
through as-is.

Usage:
    python -m src.load_excalibr_calibrations [JSON_DIR] [WORKBOOK]
"""

import json
from pathlib import Path

import click
import openpyxl

DEFAULT_JSON_DIR = Path("data/input/mave_calibration/excalibr/exc_pp_calib_final_fixedmapping_clinvar/json")
DEFAULT_WORKBOOK = Path("data/output/supplementary_data/Supplementary_Data_4.xlsx")

SHEET_NAME = "ExCALIBR_calibrations"
RANGE_POINTS = [-8, -7, -6, -5, -4, -3, -2, -1, 1, 2, 3, 4, 5, 6, 7, 8]
COLUMNS = (
    ["dataset", "prior"]
    + [f"range_{point}" for point in RANGE_POINTS]
    + ["relax", "n_c", "benign_method", "clinvar_2018", "scoreset_flipped"]
)


def format_endpoint(value):
    if value == float("-inf"):
        return "-inf"
    if value == float("inf"):
        return "inf"
    return str(value)


def format_range(point_ranges, point):
    """Return the "low high" string for one point, or None if it has no range.

    Raises ValueError if the point has more than one range entry.
    """
    entries = point_ranges.get(str(point), [])
    if len(entries) > 1:
        raise ValueError(f"point {point} has {len(entries)} range entries, expected at most 1")
    if not entries:
        return None
    low, high = entries[0]
    return f"{format_endpoint(low)} {format_endpoint(high)}"


def load_calibration_row(json_path):
    data = json.loads(json_path.read_text())
    point_ranges = data["point_ranges"]

    row = {"dataset": json_path.stem, "prior": data["prior"]}
    for point in RANGE_POINTS:
        row[f"range_{point}"] = format_range(point_ranges, point)
    row["relax"] = bool(data["relax"])
    row["n_c"] = data["n_c"]
    row["benign_method"] = data["benign_method"]
    row["clinvar_2018"] = bool(data["clinvar_2018"])
    row["scoreset_flipped"] = bool(data["scoreset_flipped"])
    return row


def load_all_calibrations(json_dir):
    json_paths = sorted(json_dir.glob("*.json"), key=lambda p: p.stem)
    return [load_calibration_row(path) for path in json_paths]


def write_calibrations_sheet(workbook_path, rows):
    workbook = openpyxl.load_workbook(workbook_path)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"{workbook_path} has no {SHEET_NAME!r} sheet")

    sheet = workbook[SHEET_NAME]
    sheet.delete_rows(1, sheet.max_row)
    sheet.append(COLUMNS)
    for row in rows:
        sheet.append([row[column] for column in COLUMNS])

    workbook.save(workbook_path)


@click.command(help=__doc__)
@click.argument(
    "json_dir",
    required=False,
    default=DEFAULT_JSON_DIR,
    type=click.Path(exists=True, file_okay=False, path_type=Path),
)
@click.argument(
    "workbook",
    required=False,
    default=DEFAULT_WORKBOOK,
    type=click.Path(exists=True, dir_okay=False, path_type=Path),
)
def main(json_dir, workbook):
    try:
        rows = load_all_calibrations(json_dir)
        write_calibrations_sheet(workbook, rows)
    except (KeyError, ValueError) as exc:
        raise click.ClickException(str(exc)) from exc

    click.echo(f"Wrote {len(rows)} row(s) to the {SHEET_NAME!r} sheet of {workbook}")


if __name__ == "__main__":
    main()
