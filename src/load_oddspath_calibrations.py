#!/usr/bin/env python3
"""Load OddsPath calibrations into Supplementary Data 4.

Reads `OddsPath_calculations.ipynb`'s output (default:
`data/output/mave_calibration/OddsPath_calibrations.csv.gz`) and writes one
row per dataset into the `OddsPath_calibrations` sheet of a workbook
(default: `data/output/supplementary_data/Supplementary_Data_4.xlsx`),
overwriting that sheet's existing contents in place. Other sheets are left
untouched.

Unlike `load_excalibr_calibrations`, no reformatting happens: the CSV
already matches the sheet's column layout, so each row is copied through
as-is (sorted by `Dataset`), with truly missing values written as blank
cells. That's not the same as the literal text "None", which several
columns can legitimately hold as a real value -- e.g. `Pseudocount Details`
defaults to "None" (not blank) when no pseudocount was needed, and
`KCNQ4_Zheng_2022_v12_homozygous`'s `Evidence Code Abnormal` is deliberately
overridden to "None" in the notebook (masked, not missing) -- so this script
reads the CSV with pandas' default NA-string recognition turned off to keep
that word intact instead of silently collapsing it to a blank cell.
`OddsNormal`/`OddsAbnormal` can also legitimately hold a string (e.g.
`"No benign controls"`) instead of a number, when a dataset lacks the
controls needed to compute a likelihood ratio -- these pass through
unchanged, same as the source CSV.

Usage:
    python -m src.load_oddspath_calibrations [CSV] [WORKBOOK]
"""

from pathlib import Path

import click
import openpyxl
import pandas as pd

DEFAULT_CSV = Path("data/output/mave_calibration/OddsPath_calibrations.csv.gz")
DEFAULT_WORKBOOK = Path("data/output/supplementary_data/Supplementary_Data_4.xlsx")

SHEET_NAME = "OddsPath_calibrations"
COLUMNS = [
    "Dataset",
    "Total Controls",
    "OddsNormal",
    "OddsAbnormal",
    "Pathogenic Controls",
    "Benign Controls",
    "Prior Probability Pathogenic",
    "Total Assay Abnormal",
    "True Path in Abnormal",
    "Total Assay Normal",
    "True Path in Normal",
    "Pseudocount Details",
    "Evidence Code Normal",
    "Evidence Code Abnormal",
]


NUMERIC_OR_STRING_COLUMNS = ["OddsNormal", "OddsAbnormal"]


def _coerce_numeric_or_string(value):
    """Parse `value` as a float if possible, otherwise return it unchanged.

    `OddsNormal`/`OddsAbnormal` mix numbers with strings like "No benign
    controls" across rows -- when a column has any non-numeric value, pandas
    reads the *whole* column as strings, turning e.g. `192.5806` into the
    string `"192.5806"` rather than a float. This restores the numeric cells
    to actual numbers cell-by-cell while leaving genuine non-numeric strings
    (which `float()` rejects) untouched.
    """
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def load_all_calibrations(csv_path):
    """Read the OddsPath calibrations CSV, sorted by `Dataset`.

    Raises ValueError if the CSV's columns don't match `COLUMNS` exactly, so
    a schema change in `OddsPath_calculations.ipynb` fails loudly instead of
    silently writing a mismatched sheet.
    """
    # keep_default_na=False/na_values=[""]: pandas' default NA sentinels
    # include the bare word "None" (also "NA", "null", etc.), but the
    # notebook writes that word deliberately as a real value -- e.g.
    # Pseudocount Details defaults to the string "None" when no pseudocount
    # was added, and KCNQ4_Zheng_2022_v12_homozygous's Evidence Code
    # Abnormal is explicitly overridden to "None" (masked, not missing; see
    # the notebook cell just above `all_vars.to_csv(...)`). Only a truly
    # empty cell should become a blank sheet cell.
    df = pd.read_csv(csv_path, keep_default_na=False, na_values=[""])
    if list(df.columns) != COLUMNS:
        raise ValueError(f"{csv_path} has columns {list(df.columns)}, expected {COLUMNS}")

    df = df.sort_values("Dataset").reset_index(drop=True)
    df = df.astype(object).where(df.notna(), None)
    for column in NUMERIC_OR_STRING_COLUMNS:
        df[column] = df[column].map(lambda v: v if v is None else _coerce_numeric_or_string(v))
    return df.to_dict("records")


def write_calibrations_sheet(workbook_path, rows):
    workbook = openpyxl.load_workbook(workbook_path)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"{workbook_path} has no {SHEET_NAME!r} sheet")

    sheet = workbook[SHEET_NAME]
    sheet.delete_rows(1, sheet.max_row)
    sheet.append(COLUMNS)
    for row in rows:
        sheet.append([row[column] for column in COLUMNS])

    workbook.save(workbook_path)


@click.command(help=__doc__)
@click.argument(
    "csv_path",
    metavar="CSV",
    required=False,
    default=DEFAULT_CSV,
    type=click.Path(exists=True, dir_okay=False, path_type=Path),
)
@click.argument(
    "workbook",
    required=False,
    default=DEFAULT_WORKBOOK,
    type=click.Path(exists=True, dir_okay=False, path_type=Path),
)
def main(csv_path, workbook):
    try:
        rows = load_all_calibrations(csv_path)
        write_calibrations_sheet(workbook, rows)
    except ValueError as exc:
        raise click.ClickException(str(exc)) from exc

    click.echo(f"Wrote {len(rows)} row(s) to the {SHEET_NAME!r} sheet of {workbook}")


if __name__ == "__main__":
    main()
