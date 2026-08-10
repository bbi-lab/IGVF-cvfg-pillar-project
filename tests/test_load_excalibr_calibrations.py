import json

import openpyxl
import pytest
from click.testing import CliRunner

from src.load_excalibr_calibrations import (
    COLUMNS,
    SHEET_NAME,
    load_all_calibrations,
    main,
    write_calibrations_sheet,
)


def _write_json(path, dataset, **overrides):
    point_ranges = {str(point): [] for point in range(-8, 9) if point != 0}
    point_ranges.update(overrides.pop("point_ranges", {}))
    data = {
        "prior": 0.25,
        "point_ranges": point_ranges,
        "dataset": dataset,
        "relax": 1,
        "n_c": "3c_avg",
        "benign_method": "avg",
        "clinvar_2018": 0,
        "scoreset_flipped": 0,
    }
    data.update(overrides)
    (path / f"{dataset}.json").write_text(json.dumps(data))


@pytest.fixture
def json_dir(tmp_path):
    json_dir = tmp_path / "json"
    json_dir.mkdir()
    _write_json(
        json_dir,
        "GENE_A",
        point_ranges={
            "1": [[-1.5, float("inf")]],
            "-1": [[float("-inf"), 0.2]],
        },
        clinvar_2018=1,
    )
    _write_json(json_dir, "GENE_B", relax=0, scoreset_flipped=1)
    return json_dir


@pytest.fixture
def workbook_path(tmp_path):
    path = tmp_path / "workbook.xlsx"
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    sheet = workbook.create_sheet(SHEET_NAME)
    sheet.append(COLUMNS)
    sheet.append(["STALE_DATASET"] + [None] * (len(COLUMNS) - 1))
    workbook.create_sheet("Other Sheet")["A1"] = "untouched"
    workbook.save(path)
    return path


def test_load_all_calibrations_formats_ranges_and_booleans(json_dir):
    rows = load_all_calibrations(json_dir)
    by_dataset = {row["dataset"]: row for row in rows}

    gene_a = by_dataset["GENE_A"]
    assert gene_a["range_1"] == "-1.5 inf"
    assert gene_a["range_-1"] == "-inf 0.2"
    assert gene_a["range_2"] is None
    assert gene_a["clinvar_2018"] is True
    assert gene_a["relax"] is True

    gene_b = by_dataset["GENE_B"]
    assert gene_b["relax"] is False
    assert gene_b["scoreset_flipped"] is True


def test_load_all_calibrations_formats_multiple_ranges(tmp_path):
    json_dir = tmp_path / "json"
    json_dir.mkdir()
    _write_json(
        json_dir,
        "GENE_C",
        point_ranges={
            "1": [[0.46, float("inf")], [-0.56, -0.52]],
            "-1": [[float("-inf"), -1.0]],
        },
    )

    rows = load_all_calibrations(json_dir)
    gene_c = {row["dataset"]: row for row in rows}["GENE_C"]

    # Sorted by lower endpoint regardless of source order.
    assert gene_c["range_1"] == "-0.56 -0.52, 0.46 inf"
    assert gene_c["range_-1"] == "-inf -1.0"


def test_write_calibrations_sheet_overwrites_only_target_sheet(json_dir, workbook_path):
    rows = load_all_calibrations(json_dir)
    write_calibrations_sheet(workbook_path, rows)

    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook[SHEET_NAME]
    values = list(sheet.iter_rows(values_only=True))
    assert values[0] == tuple(COLUMNS)
    datasets = [row[0] for row in values[1:]]
    assert datasets == ["GENE_A", "GENE_B"]
    assert "STALE_DATASET" not in datasets
    assert workbook["Other Sheet"]["A1"].value == "untouched"


def test_main_cli_reports_row_count(json_dir, workbook_path):
    runner = CliRunner()
    result = runner.invoke(main, [str(json_dir), str(workbook_path)])

    assert result.exit_code == 0
    assert "Wrote 2 row(s)" in result.output
