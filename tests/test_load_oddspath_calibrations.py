import math

import openpyxl
import pandas as pd
import pytest
from click.testing import CliRunner

from src.load_oddspath_calibrations import (
    COLUMNS,
    SHEET_NAME,
    load_all_calibrations,
    main,
    write_calibrations_sheet,
)


def _write_csv(path, rows):
    pd.DataFrame(rows, columns=COLUMNS).to_csv(path, index=False)


@pytest.fixture
def csv_path(tmp_path):
    path = tmp_path / "OddsPath_calibrations.csv.gz"
    _write_csv(
        path,
        [
            {
                # Mirrors KCNQ4_Zheng_2022_v12_homozygous: the notebook writes
                # the literal string "None" both as Pseudocount Details'
                # default (no pseudocount added) and as a deliberate,
                # explicit override of Evidence Code Abnormal (masked, not
                # missing) -- neither should collapse to a blank cell.
                "Dataset": "GENE_B",
                "Total Controls": 14,
                "OddsNormal": 2.0,
                "OddsAbnormal": "No functionally abnormal controls",
                "Pathogenic Controls": 7,
                "Benign Controls": 7,
                "Prior Probability Pathogenic": 0.5,
                "Total Assay Abnormal": 0,
                "True Path in Abnormal": 0,
                "Total Assay Normal": 140,
                "True Path in Normal": 4,
                "Pseudocount Details": "None",
                "Evidence Code Normal": "Indeterminate",
                "Evidence Code Abnormal": "None",
            },
            {
                "Dataset": "GENE_A",
                "Total Controls": 323,
                "OddsNormal": 0.0165,
                "OddsAbnormal": 192.5806,
                "Pathogenic Controls": 124,
                "Benign Controls": 199,
                "Prior Probability Pathogenic": 0.3839,
                "Total Assay Abnormal": 505,
                "True Path in Abnormal": 120,
                "Total Assay Normal": 2694,
                "True Path in Normal": 2,
                "Pseudocount Details": "Added 1 misclassified benign variant to abnormal function class",
                "Evidence Code Normal": "BS3_strong",
                "Evidence Code Abnormal": "PS3_strong",
            },
            {
                # A genuinely missing value (empty CSV cell) should still
                # become a blank sheet cell, distinct from the literal
                # string "None" above.
                "Dataset": "GENE_C",
                "Total Controls": 0,
                "OddsNormal": "No controls",
                "OddsAbnormal": "No controls",
                "Pathogenic Controls": 0,
                "Benign Controls": 0,
                "Prior Probability Pathogenic": 0,
                "Total Assay Abnormal": 0,
                "True Path in Abnormal": 0,
                "Total Assay Normal": 0,
                "True Path in Normal": 0,
                "Pseudocount Details": "No controls",
                "Evidence Code Normal": None,
                "Evidence Code Abnormal": None,
            },
        ],
    )
    return path


@pytest.fixture
def workbook_path(tmp_path):
    path = tmp_path / "workbook.xlsx"
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    sheet = workbook.create_sheet(SHEET_NAME)
    sheet.append(COLUMNS)
    sheet.append(["STALE_DATASET"] + [None] * (len(COLUMNS) - 1))
    workbook.create_sheet("Other Sheet")["A1"] = "untouched"
    workbook.save(path)
    return path


def test_load_all_calibrations_sorts_and_blanks_missing_values(csv_path):
    rows = load_all_calibrations(csv_path)

    assert [row["Dataset"] for row in rows] == ["GENE_A", "GENE_B", "GENE_C"]

    gene_a = rows[0]
    assert gene_a["OddsAbnormal"] == pytest.approx(192.5806)
    assert gene_a["Evidence Code Normal"] == "BS3_strong"

    # The literal string "None" is a real, deliberate value (see
    # KCNQ4_Zheng_2022_v12_homozygous) and must survive as text, not collapse
    # into a blank cell the way pandas' default NA-string recognition would.
    gene_b = rows[1]
    assert gene_b["OddsAbnormal"] == "No functionally abnormal controls"
    assert gene_b["Pseudocount Details"] == "None"
    assert gene_b["Evidence Code Abnormal"] == "None"

    # A genuinely empty CSV cell is still blanked out to Python None.
    gene_c = rows[2]
    assert gene_c["Evidence Code Normal"] is None
    assert gene_c["Evidence Code Abnormal"] is None


def test_load_all_calibrations_rejects_mismatched_columns(tmp_path):
    path = tmp_path / "bad.csv.gz"
    pd.DataFrame({"Dataset": ["GENE_A"], "Unexpected Column": [1]}).to_csv(path, index=False)

    with pytest.raises(ValueError, match="expected"):
        load_all_calibrations(path)


def test_write_calibrations_sheet_overwrites_only_target_sheet(csv_path, workbook_path):
    rows = load_all_calibrations(csv_path)
    write_calibrations_sheet(workbook_path, rows)

    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook[SHEET_NAME]
    values = list(sheet.iter_rows(values_only=True))
    assert values[0] == tuple(COLUMNS)
    datasets = [row[0] for row in values[1:]]
    assert datasets == ["GENE_A", "GENE_B", "GENE_C"]
    assert "STALE_DATASET" not in datasets
    assert workbook["Other Sheet"]["A1"].value == "untouched"

    gene_b_row = dict(zip(COLUMNS, values[2]))
    assert gene_b_row["Pseudocount Details"] == "None"
    assert gene_b_row["Evidence Code Abnormal"] == "None"

    gene_c_row = dict(zip(COLUMNS, values[3]))
    assert gene_c_row["Evidence Code Abnormal"] is None
    # openpyxl round-trips a written None back as None, not NaN.
    assert not isinstance(gene_c_row["Evidence Code Abnormal"], float) or not math.isnan(
        gene_c_row["Evidence Code Abnormal"]
    )


def test_main_cli_reports_row_count(csv_path, workbook_path):
    runner = CliRunner()
    result = runner.invoke(main, [str(csv_path), str(workbook_path)])

    assert result.exit_code == 0
    assert "Wrote 3 row(s)" in result.output
